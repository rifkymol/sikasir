from io import BytesIO
from typing import Optional
from datetime import datetime
import csv

import openpyxl
import pandas as pd
from fastapi import APIRouter, HTTPException, Depends, Query, status, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlmodel import Session, select

from ..auth import require_admin_actor
from ..database import get_session
from ..models import Product
from ..schemas import ProductCreate, ProductUpdate

router = APIRouter(prefix="/products", tags=["Products"])


@router.get("/import/template")
async def download_import_template(_: dict = Depends(require_admin_actor)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import Produk"

    headers = ["SKU", "Nama Produk", "Harga", "Stok", "Kategori", "Deskripsi"]
    header_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    example_data = [
        ["PRD-001", "Contoh Produk A", 15000, 50, "Makanan", "Deskripsi opsional"],
        ["PRD-002", "Contoh Produk B", 25000, 30, "Minuman", ""],
    ]
    example_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
    example_font = Font(color="6B7280", italic=True)

    for row_num, row_data in enumerate(example_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.fill = example_fill
            cell.font = example_font

    ws_guide = wb.create_sheet(title="Panduan")
    guide_data = [
        ["Kolom", "Keterangan", "Wajib?", "Contoh"],
        ["SKU", "Kode unik produk", "Ya", "PRD-001"],
        ["Nama Produk", "Nama produk yang dijual", "Ya", "Kopi Susu"],
        ["Harga", "Harga dalam Rupiah (angka saja)", "Ya", "15000"],
        ["Stok", "Jumlah stok awal (angka saja)", "Ya", "100"],
        ["Kategori", "Kategori produk", "Tidak", "Minuman"],
        ["Deskripsi", "Deskripsi singkat produk", "Tidak", "Kopi dengan susu segar"],
    ]
    guide_header_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")

    for row_num, row_data in enumerate(guide_data, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws_guide.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:
                cell.fill = guide_header_fill
                cell.font = Font(bold=True, color="FFFFFF")

    column_widths = [15, 25, 15, 10, 15, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    guide_widths = [18, 45, 10, 28]
    for i, width in enumerate(guide_widths, 1):
        ws_guide.column_dimensions[get_column_letter(i)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="template_import_produk.xlsx"'},
    )


@router.post("/import")
async def import_products(
    file: UploadFile = File(...),
    mode: str = Form(default="import"),
    _: dict = Depends(require_admin_actor),
    session: Session = Depends(get_session),
):
    filename = (file.filename or "").lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".csv")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Format file tidak didukung. Gunakan .xlsx atau .csv",
        )

    try:
        content = await file.read()
        if filename.endswith(".xlsx"):
            df = pd.read_excel(BytesIO(content), sheet_name=0)
        else:
            text_data = content.decode("utf-8-sig", errors="ignore")
            df = pd.read_csv(BytesIO(text_data.encode("utf-8")), quoting=csv.QUOTE_MINIMAL)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Gagal membaca file: {str(exc)}",
        )

    required_columns = ["SKU", "Nama Produk", "Harga", "Stok"]
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Kolom wajib tidak ditemukan: {', '.join(missing_columns)}. Pastikan menggunakan template yang benar.",
        )

    existing_skus = {sku.strip().upper() for sku in session.exec(select(Product.sku)).all() if sku}
    seen_skus = set()
    success_rows = []
    error_rows = []

    for index, row in df.iterrows():
        row_number = index + 2
        errors = []

        sku_raw = row.get("SKU")
        sku = "" if pd.isna(sku_raw) else str(sku_raw).strip()
        if not sku:
            errors.append("SKU tidak boleh kosong")

        name_raw = row.get("Nama Produk")
        name = "" if pd.isna(name_raw) else str(name_raw).strip()
        if not name:
            errors.append("Nama Produk tidak boleh kosong")

        try:
            harga = int(float(row.get("Harga")))
            if harga <= 0:
                errors.append("Harga harus lebih besar dari 0")
        except (ValueError, TypeError):
            errors.append("Harga harus berupa angka")
            harga = None

        try:
            stok = int(float(row.get("Stok")))
            if stok < 0:
                errors.append("Stok tidak boleh negatif")
        except (ValueError, TypeError):
            errors.append("Stok harus berupa angka")
            stok = None

        sku_key = sku.upper()
        if sku and sku_key in existing_skus:
            errors.append(f'SKU "{sku}" sudah ada di database')
        if sku and sku_key in seen_skus:
            errors.append(f'SKU "{sku}" duplikat di file upload')

        category_raw = row.get("Kategori")
        description_raw = row.get("Deskripsi")
        category = "General" if pd.isna(category_raw) or str(category_raw).strip() == "" else str(category_raw).strip()
        description = None if pd.isna(description_raw) or str(description_raw).strip() == "" else str(description_raw).strip()

        if errors:
            error_rows.append(
                {
                    "row": row_number,
                    "sku": sku,
                    "nama": name,
                    "errors": errors,
                }
            )
            continue

        seen_skus.add(sku_key)
        success_rows.append(
            {
                "sku": sku,
                "name": name,
                "price": harga,
                "stock": stok,
                "category": category,
                "description": description,
            }
        )

    if mode == "preview":
        return {
            "success": True,
            "preview": True,
            "total_rows": len(df),
            "valid_rows": len(success_rows),
            "error_rows": error_rows,
            "data": success_rows,
        }

    if not success_rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Tidak ada data valid yang bisa diimport.",
        )

    try:
        for item in success_rows:
            product = Product.model_validate(
                {
                    "sku": item["sku"],
                    "name": item["name"],
                    "price": float(item["price"]),
                    "stock": int(item["stock"]),
                    "category": item.get("category") or "General",
                    "description": item.get("description"),
                    "is_active": True,
                }
            )
            session.add(product)

        session.commit()

        return {
            "success": True,
            "message": f"{len(success_rows)} produk berhasil diimport.",
            "imported": len(success_rows),
            "skipped": len(error_rows),
            "error_rows": error_rows,
        }
    except Exception as exc:
        session.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Import gagal: {str(exc)}",
        )

@router.get("/")
async def get_products(
        skip: int = Query(default=0, ge=0, description="Number of records to skip"),
        limit: int = Query(default=100, le=200, description="Maximum number of records to return"),
        category: Optional[str] = Query(default=None),
        search: Optional[str] = Query(default=None, description="Search by name or SKU"),
        show_inactive: bool = Query(default=False, description="Include inactive products"),
        session: Session = Depends(get_session)
):

    query = select(Product)
    if not show_inactive:
        query = query.where(Product.is_active == True)
    if category:
        query = query.where(Product.category == category)
    if search:
        search_pattern = f"%{search}%"
        query = query.where((Product.name.ilike(search_pattern)) | (Product.sku.ilike(search_pattern)))
    products = session.exec(query.offset(skip).limit(limit)).all()
    return {"success": True, "message": "Products fetched successfully", "data": products}

@router.get("/{product_id}")
async def get_product(product_id: int, session: Session = Depends(get_session)):
    product = session.get(Product, product_id)
    if not product or not product.is_active:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Product not found"
        )
    return {"success": True, "message": "Product fetched successfully", "data": product}

@router.post("/", status_code=status.HTTP_201_CREATED)
async def create_product(product_data: ProductCreate, session: Session = Depends(get_session)):
    existing = session.exec(select(Product).where(Product.sku == product_data.sku)).first()

    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="SKU already exists"
        )

    # SQLModel + Pydantic v2: use model_validate for table models with default_factory fields.
    product = Product.model_validate(product_data.model_dump())
    session.add(product)
    session.commit()
    session.refresh(product)
    return {"success": True, "message": "Product created successfully", "data": product}

@router.put("/{product_id}")
async def update_product(product_id: int, product_data: ProductUpdate, session: Session = Depends(get_session)):
    product = session.get(Product, product_id)
    if not product or not product.is_active:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Product not found"
        )

    update_data = product_data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(product, key, value)

    product.updated_at = datetime.utcnow()

    session.add(product)
    session.commit()
    session.refresh(product)
    return {"success": True, "message": "Product updated successfully", "data": product}

@router.patch("/{product_id}/toggle")
async def toggle_product_status(product_id: int, session: Session = Depends(get_session)):
    product = session.get(Product, product_id)
    if not product:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Product not found"
        )
    product.is_active = not product.is_active
    product.updated_at = datetime.utcnow()
    session.add(product)
    session.commit()
    session.refresh(product)
    label = "diaktifkan" if product.is_active else "dinonaktifkan"
    return {"success": True, "message": f"Produk berhasil {label}", "data": product}

@router.delete("/{product_id}")
async def delete_product(product_id: int, session: Session = Depends(get_session)):
    product = session.get(Product, product_id)
    if not product:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Product not found"
        )
    session.delete(product)
    session.commit()
    return {"success": True, "message": "Produk berhasil dihapus", "data": None}